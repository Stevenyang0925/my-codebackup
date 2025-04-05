"""
Excel表格解析器模块
"""

import os
from typing import Dict, Any, List
from .base_parser import BaseParser
from ..utils.logger import get_logger
from ..utils.exceptions import FileParsingError

logger = get_logger(__name__)

class ExcelParser(BaseParser):
    """
    Excel表格解析器，用于解析.xlsx和.xls文件
    """
    
    def get_supported_extensions(self) -> List[str]:
        """
        获取支持的文件扩展名列表
        
        Returns:
            List[str]: 支持的文件扩展名列表
        """
        return ['.xlsx', '.xls']
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        解析Excel表格
        
        Args:
            file_path: 文件路径
            
        Returns:
            Dict[str, Any]: 解析结果
        """
        if not self.is_supported(file_path):
            logger.error(f"不支持的文件类型传递给 ExcelParser: {file_path}")
            raise FileParsingError(f"ExcelParser 不支持的文件类型: {file_path}")
        
        title = self.extract_title(file_path)
        content = []
        logger.info(f"开始解析 Excel 文件: {file_path}")
        
        try:
            # 尝试使用pandas
            import pandas as pd
            logger.debug(f"尝试使用 pandas 解析: {file_path}")
            
            # 读取所有工作表
            excel_file = pd.ExcelFile(file_path)
            logger.debug(f"找到工作表: {excel_file.sheet_names}")
            
            # 处理每个工作表
            for sheet_name in excel_file.sheet_names:
                logger.debug(f"正在处理工作表: {sheet_name}")
                # 添加工作表名称作为标题
                content.append({
                    "type": "heading",
                    "level": 2,
                    "content": f"工作表: {sheet_name}"
                })
                
                # 读取工作表数据
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                # 将DataFrame转换为表格数据
                table_data = [df.columns.tolist()]  # 表头
                table_data.extend(df.values.tolist())  # 数据行
                
                # 添加表格
                content.append({
                    "type": "table",
                    "content": "",
                    "additional_info": table_data
                })
            
            logger.info(f"使用 pandas 成功解析 Excel 文件: {file_path}")
            return {
                "title": title,
                "content": content
            }
            
        except ImportError:
            logger.warning("未找到 pandas 库，尝试使用 openpyxl 解析 Excel 文件。")
            try:
                # 如果pandas不可用，尝试使用openpyxl
                import openpyxl
                logger.debug(f"尝试使用 openpyxl 解析: {file_path}")
                
                # 打开工作簿
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                logger.debug(f"找到工作表: {workbook.sheetnames}")
                
                # 处理每个工作表
                for sheet_name in workbook.sheetnames:
                    logger.debug(f"正在处理工作表: {sheet_name}")
                    sheet = workbook[sheet_name]
                    
                    # 添加工作表名称作为标题
                    content.append({
                        "type": "heading",
                        "level": 2,
                        "content": f"工作表: {sheet_name}"
                    })
                    
                    # 提取表格数据
                    table_data = []
                    for row in sheet.rows:
                        row_data = [cell.value for cell in row]
                        table_data.append(row_data)
                    
                    # 添加表格
                    if table_data:
                        content.append({
                            "type": "table",
                            "content": "",
                            "additional_info": table_data
                        })
                
                logger.info(f"使用 openpyxl 成功解析 Excel 文件: {file_path}")
                return {
                    "title": title,
                    "content": content
                }
                
            except ImportError:
                logger.error("无法解析Excel文件：pandas 和 openpyxl 库均未安装。")
                return {
                    "title": title,
                    "content": [{
                        "type": "text",
                        "content": "无法解析Excel文件，请安装pandas或openpyxl库: pip install pandas openpyxl"
                    }]
                }
            except Exception as e_openpyxl:
                logger.exception(f"使用 openpyxl 解析 Excel 文件时发生错误: {file_path}", exc_info=True)
                raise FileParsingError(f"使用 openpyxl 解析 Excel 文件时出错: {e_openpyxl}") from e_openpyxl
        except Exception as e_pandas:
            logger.exception(f"使用 pandas 解析 Excel 文件时发生错误: {file_path}", exc_info=True)
            raise FileParsingError(f"使用 pandas 解析 Excel 文件时出错: {e_pandas}") from e_pandas